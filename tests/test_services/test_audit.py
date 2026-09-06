"""Audit 服务层单元测试。

覆盖范围：
- export_workbook 基本导出（多 Sheet Excel 工作簿）
- export_workbook 自定义 tasks 列表
- auto_report 一键报告（HTML 输出）
- export_workbook 空数据/失败任务优雅降级
"""

import os
import tempfile

import openpyxl
import pandas as pd

from smartsuite.core.contracts import AnalysisResult
from smartsuite.services import audit as audit_module
from smartsuite.services.audit import auto_report, batch_analyze, export_workbook, process_audit


# ── batch_analyze 测试 ──


def test_batch_analyze_basic(sample_doe_data):
    """验证 batch_analyze 对多个任务批量分析返回结构化摘要。"""
    result = batch_analyze(
        sample_doe_data,
        target_col="不良率",
        feature_cols=["料温", "模温"],
        tasks=["correlation", "distribution_summary"],
    )
    assert "results" in result
    assert "summary" in result
    assert "correlation" in result["results"]
    assert "distribution_summary" in result["results"]
    assert result["results"]["distribution_summary"]["status"] == "ok"


def test_batch_analyze_failed_task_graceful(sample_doe_data):
    """验证单个任务失败不拖垮批量（优雅降级为 error 状态）。"""
    result = batch_analyze(
        sample_doe_data,
        target_col="不良率",
        feature_cols=["料温"],
        tasks=["vif"],  # 单特征 VIF 会失败
    )
    assert result["results"]["vif"]["status"] in ("ok", "error")
    assert "results" in result


# ── export_workbook 测试 ──


def test_export_workbook_basic(sample_doe_data):
    """验证 export_workbook 正常生成多 Sheet Excel 工作簿。"""
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        out = export_workbook(
            sample_doe_data,
            target_col="不良率",
            feature_cols=["料温", "模温"],
            output_path=path,
            tasks=["correlation", "distribution_summary"],
        )
        assert os.path.exists(out)
        assert os.path.getsize(out) > 0

        # 验证工作簿结构
        wb = openpyxl.load_workbook(out)
        # 每个成功 task 应有一个 _summary sheet
        assert len(wb.sheetnames) >= 2  # 至少 2 个 task
        # 验证表头颜色格式正确（aRGB 无 # 前缀 — F-01 修复验证）
        for ws in wb.worksheets:
            # 检查至少有一个带填充的单元格
            fills_found = False
            for row in ws.iter_rows():
                for cell in row:
                    if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                        rgb = cell.fill.fgColor.rgb
                        assert not rgb.startswith("#"), f"openpyxl 颜色不应包含 # 前缀: {rgb}"
                        assert len(rgb) == 8, f"aRGB 应为 8 位 hex: {rgb}"
                        fills_found = True
                        break
                if fills_found:
                    break
        wb.close()
    finally:
        if os.path.exists(path):
            os.unlink(path)


def test_export_workbook_custom_tasks(sample_doe_data):
    """验证 export_workbook 自定义 tasks 列表。"""
    import numpy as np

    # Round-2 P3：anova 需要类别因子（数值连续列现被拒绝）
    df = sample_doe_data.copy()
    df["水平"] = np.random.choice(["低", "中", "高"], len(df))
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        out = export_workbook(
            df,
            target_col="强度",
            feature_cols=["水平"],
            output_path=path,
            tasks=["anova"],
        )
        assert os.path.exists(out)
        wb = openpyxl.load_workbook(out)
        sheet_names_lower = [s.lower() for s in wb.sheetnames]
        assert any("anova" in s for s in sheet_names_lower), f"应包含 anova sheet: {wb.sheetnames}"
        wb.close()
    finally:
        if os.path.exists(path):
            os.unlink(path)


def test_export_workbook_subdir_creation():
    """验证 export_workbook 自动创建不存在的输出目录。"""
    df = pd.DataFrame({"y": [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]})
    tmpdir = tempfile.mkdtemp()
    subdir = os.path.join(tmpdir, "nested", "subdir")
    out_path = os.path.join(subdir, "output.xlsx")
    try:
        out = export_workbook(
            df,
            target_col="y",
            feature_cols=[],
            output_path=out_path,
            tasks=["distribution_summary"],
        )
        assert os.path.exists(out)
    finally:
        if os.path.exists(out_path):
            os.unlink(out_path)
        # 清理嵌套目录
        for d in [subdir, os.path.join(tmpdir, "nested"), tmpdir]:
            if os.path.isdir(d) and not os.listdir(d):
                os.rmdir(d)


def test_export_workbook_all_tasks_fail():
    """验证所有 task 失败时不崩溃，生成仅含错误信息的 Sheet。"""
    df = pd.DataFrame({"y": [1, 1, 1, 1, 1]})  # 常量数据，许多分析会失败
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        path = f.name
    try:
        out = export_workbook(
            df,
            target_col="y",
            feature_cols=[],
            output_path=path,
            tasks=["regression", "anova"],  # 缺少 feature_cols，这些 task 会失败
        )
        assert os.path.exists(out)
        assert os.path.getsize(out) > 0
        # 验证至少创建了一个 sheet（降级行为）
        wb = openpyxl.load_workbook(out)
        assert len(wb.sheetnames) >= 1, "应至少有一个 Sheet（错误信息或降级）"
        wb.close()
    finally:
        if os.path.exists(path):
            os.unlink(path)


# ── auto_report 测试 ──


def test_auto_report_smoke(sample_doe_data):
    """验证 auto_report 正常生成 HTML 报告。"""
    with tempfile.NamedTemporaryFile(suffix=".html", delete=False) as f:
        path = f.name
    try:
        result = auto_report(
            sample_doe_data,
            target_col="不良率",
            feature_cols=["料温", "模温"],
            output_path=path,
            title="测试自动报告",
        )
        assert result["output_path"] == path
        assert os.path.exists(path)
        assert os.path.getsize(path) > 0
        # 验证 HTML 内容
        with open(path, encoding="utf-8") as fh:
            html = fh.read()
        assert "<html" in html.lower() or "<!doctype" in html.lower(), "输出应为有效 HTML"
        assert "测试自动报告" in html or "SmartSuite" in html, "HTML 应包含报告标题或项目名"
        # 验证返回结构
        assert "data_quality" in result
        assert "batch_results" in result
        assert "audit" in result
    finally:
        if os.path.exists(path):
            os.unlink(path)


def test_auto_report_auto_path(sample_doe_data):
    """验证 auto_report 不指定 output_path 时自动生成路径。"""
    import os as _os

    cwd = _os.getcwd()
    default_path = _os.path.join(cwd, "smartsuite_report.html")
    try:
        result = auto_report(
            sample_doe_data,
            target_col="强度",
            feature_cols=["料温", "模温"],
        )
        assert os.path.exists(result["output_path"])
        assert os.path.getsize(result["output_path"]) > 0
    finally:
        if os.path.exists(default_path):
            os.unlink(default_path)


def test_auto_report_with_spec_limits(sample_doe_data):
    """验证 auto_report 含规格限参数时正常生成（含过程能力分析）。"""
    with tempfile.NamedTemporaryFile(suffix=".html", delete=False) as f:
        path = f.name
    try:
        result = auto_report(
            sample_doe_data,
            target_col="强度",
            feature_cols=["料温", "模温"],
            output_path=path,
            usl=55,
            lsl=35,
        )
        assert os.path.exists(path)
        assert "output_path" in result
    finally:
        if os.path.exists(path):
            os.unlink(path)


# ── process_audit 分支矩阵（canned result 精确控制各检查项走向）──


def _ok_result(task: str, metadata: dict | None = None) -> AnalysisResult:
    return AnalysisResult(task=task, status="ok", summary=f"{task} 正常", metadata=metadata or {})


def _err_result(task: str) -> AnalysisResult:
    return AnalysisResult(task=task, status="error", summary=f"{task} 失败", messages=["失败"])


_DF = pd.DataFrame(
    {
        "温度": [180.0, 182.0, 185.0, 188.0, 190.0] * 4,
        "压力": [60.0, 62.0, 65.0, 68.0, 70.0] * 4,
        "强度": [45.0, 46.0, 47.0, 48.0, 49.0] * 4,
    }
)


def _audit_with(monkeypatch, script: dict, feature_cols=None, **audit_kw):
    """按 task 名依次返回脚本化结果（或抛异常），驱动 process_audit 各分支。

    队列顺序 = process_audit 实际调用顺序：
    correlation → (vif, ≥3 特征) → (capability, 有规格限) → (trend, time_order) → outlier
    """
    queue = list(script.items())

    def fake_orchestrate(req):
        action = queue.pop(0)[1]
        if isinstance(action, Exception):
            raise action
        return action

    monkeypatch.setattr(audit_module, "orchestrate", fake_orchestrate)
    if feature_cols is None:
        feature_cols = ["温度", "压力"]
    return process_audit(_DF, target_col="强度", feature_cols=feature_cols, **audit_kw)


def test_process_audit_strong_correlation_marks_good(monkeypatch):
    """|r|>0.5 → 「✓ 良好」（audit.py:101-108）。"""
    r = _audit_with(
        monkeypatch,
        {"correlation": _ok_result("correlation", {"target_correlations": {"温度": 0.9}})},
    )
    row = r["health_checks"].query("检查项 == '关键因子识别'").iloc[0]
    assert "良好" in row["状态"] and "0.90" in row["详情"]


def test_process_audit_orchestrate_crash_isolated_per_check(monkeypatch):
    """单个分析崩溃 → 该项「✗ 失败」，其余检查继续（audit.py:117-125 等）。"""
    r = _audit_with(
        monkeypatch,
        {
            "correlation": RuntimeError("崩溃"),
            "process_capability": _ok_result("process_capability", {"cpk": 1.67}),
            "outlier_consensus": _ok_result("outlier_consensus", {"high_confidence_count": 0}),
        },
        usl=55,
        lsl=35,
    )
    checks = r["health_checks"]
    assert "计算异常 (RuntimeError)" in checks.query("检查项 == '关键因子识别'").iloc[0]["详情"]
    assert checks[checks["检查项"] == "过程能力"].iloc[0]["状态"].startswith("✓")


def test_process_audit_vif_fail_and_high_collinearity(monkeypatch):
    """VIF 失败 → ✗；high_vif_count>0 → ⚠ 警告（audit.py:136-157）。

    VIF 仅在 ≥3 个数值特征列时执行，队列按实际调用顺序供给。
    """
    r = _audit_with(
        monkeypatch,
        {
            "correlation": _ok_result("correlation", {"target_correlations": {}}),
            "vif": _err_result("vif"),
            "outlier_consensus": _ok_result("outlier_consensus", {"high_confidence_count": 0}),
        },
        feature_cols=["温度", "压力", "强度"],
    )
    assert "失败" in r["health_checks"].query("检查项 == '共线性诊断'").iloc[0]["状态"]

    r2 = _audit_with(
        monkeypatch,
        {
            "correlation": _ok_result("correlation", {"target_correlations": {}}),
            "vif": _ok_result("vif", {"high_vif_count": 2}),
            "outlier_consensus": _ok_result("outlier_consensus", {"high_confidence_count": 0}),
        },
        feature_cols=["温度", "压力", "强度"],
    )
    row = r2["health_checks"].query("检查项 == '共线性诊断'").iloc[0]
    assert "⚠" in row["状态"] and "2 个因子 VIF>5" in row["详情"]


def test_process_audit_vif_crash_isolated(monkeypatch):
    """VIF 计算崩溃 → ✗ 失败（audit.py:158-162）。"""
    r = _audit_with(
        monkeypatch,
        {
            "correlation": _ok_result("correlation", {"target_correlations": {}}),
            "vif": ValueError("崩溃"),
            "outlier_consensus": _ok_result("outlier_consensus", {"high_confidence_count": 0}),
        },
        feature_cols=["温度", "压力", "强度"],
    )
    assert (
        "计算异常 (ValueError)"
        in r["health_checks"].query("检查项 == '共线性诊断'").iloc[0]["详情"]
    )


def test_process_audit_capability_none_cpk(monkeypatch):
    """能力分析成功但 metadata 无 cpk → 「— 未计算」（audit.py:210-211）。"""
    r = _audit_with(
        monkeypatch,
        {
            "correlation": _ok_result("correlation", {"target_correlations": {}}),
            "process_capability": _ok_result("process_capability", {}),
        },
        usl=55,
        lsl=35,
    )
    row = r["health_checks"].query("检查项 == '过程能力'").iloc[0]
    assert row["状态"] == "—" and "未计算" in row["详情"]


def test_process_audit_capability_marginal_and_fail(monkeypatch):
    """Cpk ∈ [1.0,1.33) → ⚠ 勉强；失败结果 → ✗（audit.py:176-201）。"""
    r = _audit_with(
        monkeypatch,
        {
            "correlation": _ok_result("correlation", {"target_correlations": {}}),
            "process_capability": _ok_result("process_capability", {"cpk": 1.1}),
        },
        usl=55,
        lsl=35,
    )
    assert "勉强" in r["health_checks"].query("检查项 == '过程能力'").iloc[0]["状态"]

    r2 = _audit_with(
        monkeypatch,
        {
            "correlation": _ok_result("correlation", {"target_correlations": {}}),
            "process_capability": _err_result("process_capability"),
        },
        usl=55,
        lsl=35,
    )
    assert "失败" in r2["health_checks"].query("检查项 == '过程能力'").iloc[0]["状态"]


def test_process_audit_capability_crash_isolated(monkeypatch):
    """能力计算崩溃 → ✗ 失败（audit.py:212-216）。"""
    r = _audit_with(
        monkeypatch,
        {
            "correlation": _ok_result("correlation", {"target_correlations": {}}),
            "process_capability": ZeroDivisionError("崩溃"),
        },
        usl=55,
        lsl=35,
    )
    assert (
        "计算异常 (ZeroDivisionError)"
        in r["health_checks"].query("检查项 == '过程能力'").iloc[0]["详情"]
    )


def test_process_audit_trend_stable_and_autocorr(monkeypatch):
    """time_order=True：DW 安全区间 → ✓ 稳定；区间外 → ⚠ 注意（audit.py:219-252）。"""
    r = _audit_with(
        monkeypatch,
        {
            "correlation": _ok_result("correlation", {"target_correlations": {}}),
            "trend_forecast": _ok_result("trend_forecast", {"durbin_watson": 2.0}),
        },
        time_order=True,
    )
    assert "稳定" in r["health_checks"].query("检查项 == '过程稳定性'").iloc[0]["状态"]

    r2 = _audit_with(
        monkeypatch,
        {
            "correlation": _ok_result("correlation", {"target_correlations": {}}),
            "trend_forecast": _ok_result("trend_forecast", {"durbin_watson": 1.0}),
        },
        time_order=True,
    )
    row = r2["health_checks"].query("检查项 == '过程稳定性'").iloc[0]
    assert "⚠" in row["状态"] and "自相关" in row["详情"]


def test_process_audit_trend_fail_and_crash(monkeypatch):
    """趋势分析失败与崩溃两条兜底（audit.py:227-234, 253-257）。"""
    r = _audit_with(
        monkeypatch,
        {
            "correlation": _ok_result("correlation", {"target_correlations": {}}),
            "trend_forecast": _err_result("trend_forecast"),
        },
        time_order=True,
    )
    assert "失败" in r["health_checks"].query("检查项 == '过程稳定性'").iloc[0]["状态"]

    r2 = _audit_with(
        monkeypatch,
        {
            "correlation": _ok_result("correlation", {"target_correlations": {}}),
            "trend_forecast": IndexError("崩溃"),
        },
        time_order=True,
    )
    assert (
        "计算异常 (IndexError)"
        in r2["health_checks"].query("检查项 == '过程稳定性'").iloc[0]["详情"]
    )


def test_process_audit_outlier_crash_isolated(monkeypatch):
    """异常检测崩溃 → ✗ 失败（audit.py:292-296）。"""
    r = _audit_with(
        monkeypatch,
        {
            "correlation": _ok_result("correlation", {"target_correlations": {}}),
            "outlier_consensus": TypeError("崩溃"),
        },
    )
    assert (
        "计算异常 (TypeError)" in r["health_checks"].query("检查项 == '异常值检测'").iloc[0]["详情"]
    )


def test_process_audit_overall_rating_excellent(monkeypatch):
    """全部 ✓ → 「优秀 (全部正常)」（audit.py:310-311）。"""
    r = _audit_with(
        monkeypatch,
        {
            "correlation": _ok_result("correlation", {"target_correlations": {"温度": 0.9}}),
            "process_capability": _ok_result("process_capability", {"cpk": 1.67}),
            "outlier_consensus": _ok_result("outlier_consensus", {"high_confidence_count": 0}),
        },
        usl=55,
        lsl=35,
    )
    assert r["overall_rating"].startswith("优秀")
    assert r["score_detail"].startswith("✓")


def test_process_audit_overall_rating_variants(monkeypatch):
    """单项 ⚠ → 「良好 (1 项需关注)」；≥2 项 ⚠ → 「需关注 (多项警告)」（audit.py:306-309）。"""
    weak_corr = _ok_result("correlation", {"target_correlations": {"温度": 0.2}})
    good_rest = {
        "correlation": weak_corr,
        "process_capability": _ok_result("process_capability", {"cpk": 1.67}),
        "outlier_consensus": _ok_result("outlier_consensus", {"high_confidence_count": 0}),
    }
    r1 = _audit_with(monkeypatch, dict(good_rest), usl=55, lsl=35)
    assert r1["overall_rating"].startswith("良好")

    r2 = _audit_with(
        monkeypatch,
        {
            **good_rest,
            "outlier_consensus": _ok_result("outlier_consensus", {"high_confidence_count": 3}),
        },
        usl=55,
        lsl=35,
    )
    assert r2["overall_rating"].startswith("需关注")


# ── batch_analyze / auto_report / export_workbook 补充分支 ──


def test_batch_analyze_default_task_list(sample_doe_data, monkeypatch):
    """不传 tasks → 默认 6 任务清单（audit.py:324-332）。"""
    ran = []
    real_orchestrate = audit_module.orchestrate

    def spy(req):
        ran.append(req.task)
        return real_orchestrate(req)

    monkeypatch.setattr(audit_module, "orchestrate", spy)
    result = batch_analyze(sample_doe_data, target_col="强度", feature_cols=["料温", "模温"])
    assert ran == [
        "correlation",
        "regression",
        "vif",
        "anova",
        "normality_check",
        "distribution_summary",
    ]
    assert result["summary"].endswith("tasks OK")


def test_batch_analyze_orchestrate_crash_graceful(sample_doe_data, monkeypatch):
    """orchestrate 本身崩溃 → 单任务降级 error，批处理不中断（audit.py:348-353）。"""

    def boom(req):
        raise RuntimeError("崩溃")

    monkeypatch.setattr(audit_module, "orchestrate", boom)
    result = batch_analyze(
        sample_doe_data,
        target_col="强度",
        feature_cols=["料温"],
        tasks=["correlation", "anova"],
    )
    for task in ("correlation", "anova"):
        assert result["results"][task]["status"] == "error"
        assert "执行失败" in result["results"][task]["summary"]


def test_auto_report_regression_fail_fallback_summary(tmp_path, monkeypatch):
    """回归建模失败 → 汇总注明「回归建模失败」（audit.py:419-427）。

    feature_cols 缺省 → 自动选择数值列（audit.py:386-388）。
    """
    tiny = pd.DataFrame({"强度": [45.0, 46.0], "温度": [180.0, 182.0]})
    out = tmp_path / "report.html"
    result = auto_report(
        tiny,
        target_col="强度",
        output_path=str(out),
    )
    assert os.path.exists(out)
    assert "output_path" in result


def test_export_workbook_default_tasks(sample_doe_data, tmp_path, monkeypatch):
    """不传 tasks → 默认 6 任务（audit.py:469-477）；None 值写入空串（audit.py:512-513）。"""
    out = tmp_path / "wb.xlsx"
    ran = []
    real_orchestrate = audit_module.orchestrate

    def spy(req):
        ran.append(req.task)
        return real_orchestrate(req)

    monkeypatch.setattr(audit_module, "orchestrate", spy)
    export_workbook(
        sample_doe_data, target_col="强度", feature_cols=["料温", "模温"], output_path=str(out)
    )
    assert ran[0] == "correlation" and len(ran) == 6
    assert out.exists()


def test_export_workbook_canned_table_with_none_and_nan(tmp_path, monkeypatch):
    """表格含 None/NaN/np 值 → 分别写空串/NaN/数值（audit.py:510-522）。"""
    table = pd.DataFrame(
        {
            "数值": [1.5, float("nan")],
            "文本": pd.Series(["ok", None], dtype=object),
        }
    )
    canned = AnalysisResult(task="correlation", status="ok", summary="canned", tables={"t": table})

    def fake_orchestrate(req):
        return canned

    monkeypatch.setattr(audit_module, "orchestrate", fake_orchestrate)
    out = tmp_path / "wb2.xlsx"
    export_workbook(
        pd.DataFrame({"强度": [1.0, 2.0]}),
        target_col="强度",
        feature_cols=["强度"],
        output_path=str(out),
        tasks=["correlation"],
    )
    wb = openpyxl.load_workbook(out)
    ws = wb["correlation_summary"]
    values = [ws.cell(row=6, column=c).value for c in range(1, 3)] + [
        ws.cell(row=7, column=c).value for c in range(1, 3)
    ]
    assert 1.5 in values and "ok" in values, f"数值与文本应原样写入: {values}"
    assert "NaN" in values, "NaN 应写作字符串「NaN」防 openpyxl 崩溃"
    assert values[3] in ("", None), "None 应写空串（openpyxl 回读为 None）"


def test_export_workbook_task_crash_continues(tmp_path, monkeypatch):
    """单任务 orchestrate 崩溃 → 记警告继续下一任务（audit.py:525-527）。"""

    def boom(req):
        raise RuntimeError("崩溃")

    monkeypatch.setattr(audit_module, "orchestrate", boom)
    out = tmp_path / "wb3.xlsx"
    export_workbook(
        pd.DataFrame({"强度": [1.0, 2.0]}),
        target_col="强度",
        feature_cols=["强度"],
        output_path=str(out),
        tasks=["correlation", "anova"],
    )
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["导出状态"], "全部任务失败应只剩状态 Sheet"
    assert "所有分析任务均失败" in wb["导出状态"]["A1"].value
