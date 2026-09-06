"""Web app.py 路由直测（Flask test client，覆盖评估报告 P2-5：app.py 57% 洼地）。

覆盖面：index / csrf-token / tasks 路由、CSRF 403 分支、上传全部分支
（无文件/无扩展名/坏扩展名/GBK 编码/垃圾字节/坏 zip/zip 炸弹/非 Excel zip/
列数超限/大文件内存警告/parquet 保存失败/旧文件替换/定期清理）、analyze
全部 400 校验分支、NO_DATA 任务正路径、数据过期 400、ValidationError 400、
意外异常 500、main() debug 安全绑定。

不覆盖（导入期/环境分支，单测不可达）：app.py:21-28 flask ImportError
退出分支、129-154 SECRET_KEY 文件/env 分支（模块导入时执行）、402-409
`__main__` argparse 守卫、275-278（xlsx >100k 行需巨型 fixture，CSV 路径
已在 test_upload_limits.py 覆盖同判据）、249-250（latin-1 恒可解码，不可达）。
"""

import io
import logging
import os
import sys
import tempfile
import time
import zipfile

import pandas as pd
import pytest

from smartsuite.services.orchestrator import TASK_REGISTRY
from smartsuite.web import app as app_module
from smartsuite.web.app import app as flask_app


@pytest.fixture()
def client():
    flask_app.config.update(TESTING=True)
    with flask_app.test_client() as c:
        yield c


@pytest.fixture(autouse=True)
def _reset_upload_tracking():
    """隔离模块级临时文件追踪列表：先清理上一测试遗留文件，再清空追踪表。"""
    app_module._cleanup_uploads()
    with app_module._upload_lock:
        app_module._UPLOAD_FILES.clear()
    yield
    app_module._cleanup_uploads()
    with app_module._upload_lock:
        app_module._UPLOAD_FILES.clear()


def _csrf(client):
    resp = client.get("/api/csrf-token")
    assert resp.status_code == 200
    return resp.get_json()["token"]


def _post_csv(client, content: bytes, filename: str = "data.csv"):
    return client.post(
        "/api/upload",
        data={"file": (io.BytesIO(content), filename)},
        content_type="multipart/form-data",
        headers={"X-CSRF-Token": _csrf(client)},
    )


# ── 基础路由 ──


def test_index_route_renders(client):
    """GET / 渲染首页并为会话生成 CSRF token（app.py:186-193）。"""
    resp = client.get("/")
    assert resp.status_code == 200
    assert b"SmartSuite" in resp.data
    # index 生成的 token 与 /api/csrf-token 返回一致（已有 token 不再重新生成）
    token = client.get("/api/csrf-token").get_json()["token"]
    assert token, "访问首页后应已有 CSRF token"


def test_api_tasks_lists_all_registered(client):
    """GET /api/tasks 返回注册表全量任务+标签+分组（app.py:378-382）。"""
    resp = client.get("/api/tasks")
    assert resp.status_code == 200
    body = resp.get_json()
    assert set(body["tasks"]) == set(TASK_REGISTRY.keys())
    assert set(body["labels"]) == set(TASK_REGISTRY.keys())
    assert body["groups"], "分组不应为空"


def test_post_without_csrf_rejected_403(client):
    """无 CSRF 头的 POST 一律 403（app.py:113-125 安全分支）。"""
    resp = client.post("/api/analyze", json={"task": "anova"})
    assert resp.status_code == 403
    assert "CSRF" in resp.get_json()["error"]


# ── 上传分支 ──


def test_upload_no_file_400(client):
    """无 file 字段 → 400「请选择文件」（app.py:210-211）。"""
    resp = client.post(
        "/api/upload",
        data={},
        content_type="multipart/form-data",
        headers={"X-CSRF-Token": _csrf(client)},
    )
    assert resp.status_code == 400
    assert "请选择文件" in resp.get_json()["error"]


def test_upload_no_extension_400(client):
    resp = _post_csv(client, b"a\n1\n", filename="datafile")
    assert resp.status_code == 400
    assert "无法识别文件类型" in resp.get_json()["error"]


def test_upload_bad_extension_400(client):
    resp = _post_csv(client, b"a\n1\n", filename="data.txt")
    assert resp.status_code == 400
    assert "不支持的文件格式" in resp.get_json()["error"]


def test_upload_csv_gbk_decoded(client):
    """GBK 中文表头 CSV：utf-8 解码失败后回退 gbk 成功（app.py:233-243）。"""
    content = "强度,温度\n45.1,180\n46.3,182\n".encode("gbk")
    resp = _post_csv(client, content)
    assert resp.status_code == 200, resp.get_json()
    names = [c["name"] for c in resp.get_json()["columns"]]
    assert "强度" in names and "温度" in names


def test_upload_csv_garbage_parse_error_400(client):
    """字段数不一致的垃圾 CSV：utf-8 可解码但解析异常 → 400（app.py:244-248）。"""
    resp = _post_csv(client, b"a,b\n1,2\n1,2,3\n")
    assert resp.status_code == 400
    assert "无法解析 CSV" in resp.get_json()["error"]


def test_upload_excel_bad_zip_400(client):
    resp = _post_csv(client, b"this is not a zip", filename="data.xlsx")
    assert resp.status_code == 400
    assert "不是有效的 Excel" in resp.get_json()["error"]


def test_upload_excel_zip_bomb_400(client):
    """条目数 >1000 的 zip → 400「过多条目」（app.py:258-259）。"""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_STORED) as zf:
        for i in range(1001):
            zf.writestr(f"e{i}.xml", "x")
    resp = _post_csv(client, buf.getvalue(), filename="bomb.xlsx")
    assert resp.status_code == 400
    assert "过多条目" in resp.get_json()["error"]


def test_upload_excel_valid_zip_but_not_excel_400(client):
    """通过 zip 校验但 openpyxl 无法解析 → 400（app.py:263-267）。"""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("readme.txt", "不是 Excel 内容")
    resp = _post_csv(client, buf.getvalue(), filename="fake.xlsx")
    assert resp.status_code == 400
    assert "无法解析 Excel" in resp.get_json()["error"]


def test_upload_excel_too_many_cols_400(client):
    """501 列 xlsx → 400 列数超限（app.py:279-282）。"""
    import openpyxl

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.append([f"c{i}" for i in range(501)])
    ws.append(list(range(501)))
    buf = io.BytesIO()
    wb.save(buf)
    resp = _post_csv(client, buf.getvalue(), filename="wide.xlsx")
    assert resp.status_code == 400
    assert "超过限制" in resp.get_json()["error"]
    assert "列数" in resp.get_json()["error"]


def test_upload_large_file_logs_memory_warning(client, caplog):
    """>20MB 上传触发内存警告日志（app.py:285-287），数据本身合法。"""
    row = "1." + "2" * 208  # ≈210 字节/行 × 100k 行 ≈ 21MB > 20MiB
    buf = io.BytesIO()
    buf.write(b"v\n")
    for _ in range(100_000):
        buf.write(f"{row}\n".encode())
    with caplog.at_level(logging.WARNING, logger="smartsuite.web.app"):
        resp = _post_csv(client, buf.getvalue())
    assert resp.status_code == 200
    assert any("内存占用" in r.message for r in caplog.records)


def test_upload_parquet_save_failure_500(client, monkeypatch):
    """parquet 保存失败 → 500 中文兜底，不泄漏 traceback（app.py:295-301）。"""

    def _boom(self, *args, **kwargs):
        raise OSError("disk full")

    monkeypatch.setattr(pd.DataFrame, "to_parquet", _boom)
    resp = _post_csv(client, b"a,b\n1,2\n")
    assert resp.status_code == 500
    assert "数据保存失败" in resp.get_json()["error"]


def test_upload_replaces_old_session_file(client):
    """二次上传：旧临时文件被删除并移出追踪表（app.py:303-312）。"""
    assert _post_csv(client, b"a,b\n1,2\n").status_code == 200
    with client.session_transaction() as sess:
        old_path = sess["_data_path"]
    assert os.path.exists(old_path)
    assert _post_csv(client, b"a,b\n3,4\n").status_code == 200
    assert not os.path.exists(old_path), "旧上传文件应被删除"
    assert old_path not in app_module._UPLOAD_FILES


def test_cleanup_uploads_removes_tracked_files(client):
    """atexit 兜底清理：_cleanup_uploads 删除全部受追踪临时文件（app.py:52-60）。

    注：该函数只负责删文件，不摘除追踪表条目（摘除是 _periodic_cleanup 的职责）。
    """
    assert _post_csv(client, b"a,b\n1,2\n").status_code == 200
    with client.session_transaction() as sess:
        path = sess["_data_path"]
    assert os.path.exists(path)
    app_module._cleanup_uploads()
    assert not os.path.exists(path)


def test_periodic_cleanup_purges_stale_and_missing(client, monkeypatch):
    """定期清理：过期文件（mtime>24h）删除、幽灵路径移出追踪表（app.py:70-100）。"""
    monkeypatch.setattr(app_module, "_CLEANUP_INTERVAL", 1)  # 每次请求都触发
    stale = tempfile.NamedTemporaryFile(suffix=".parquet", delete=False)  # noqa: SIM115
    stale.close()
    old_t = time.time() - 90_000
    os.utime(stale.name, (old_t, old_t))
    ghost = stale.name + ".ghost"
    with app_module._upload_lock:
        app_module._UPLOAD_FILES.extend([stale.name, ghost])
    assert _post_csv(client, b"a,b\n1,2\n").status_code == 200
    assert not os.path.exists(stale.name), "过期文件应被定期清理删除"
    with app_module._upload_lock:
        assert stale.name not in app_module._UPLOAD_FILES
        assert ghost not in app_module._UPLOAD_FILES
        assert len(app_module._UPLOAD_FILES) == 1, "仅剩本次上传的新文件"


# ── analyze 校验分支 ──


def _analyze(client, **overrides):
    body = {"task": "anova", "targets": ["强度"], "features": ["温度"]}
    body.update(overrides)
    return client.post("/api/analyze", json=body, headers={"X-CSRF-Token": _csrf(client)})


def test_analyze_task_not_string_400(client):
    resp = _analyze(client, task=["anova"])
    assert resp.status_code == 400
    assert "task 必须是字符串" in resp.get_json()["error"]


def test_analyze_missing_task_400(client):
    resp = _analyze(client, task="")
    assert resp.status_code == 400
    assert "缺少分析任务" in resp.get_json()["error"]


def test_analyze_targets_not_list_400(client):
    resp = _analyze(client, targets="强度")
    assert resp.status_code == 400
    assert "targets 必须是字符串列表" in resp.get_json()["error"]


def test_analyze_features_not_list_400(client):
    resp = _analyze(client, features="温度")
    assert resp.status_code == 400
    assert "features 必须是字符串列表" in resp.get_json()["error"]


def test_analyze_categoricals_not_list_400(client):
    resp = _analyze(client, categoricals="工艺")
    assert resp.status_code == 400
    assert "categoricals 必须是字符串列表" in resp.get_json()["error"]


def test_analyze_params_not_dict_400(client):
    resp = _analyze(client, params=["x"])
    assert resp.status_code == 400
    assert "params 必须是字典" in resp.get_json()["error"]


def test_analyze_too_many_features_400(client):
    resp = _analyze(client, features=[f"f{i}" for i in range(101)])
    assert resp.status_code == 400
    assert "特征列数量" in resp.get_json()["error"]


def test_analyze_unknown_task_400(client):
    resp = _analyze(client, task="no_such_method")
    assert resp.status_code == 400
    assert "未知的分析任务" in resp.get_json()["error"]


def test_analyze_without_upload_400(client):
    resp = _analyze(client)
    assert resp.status_code == 400
    assert "请先上传数据文件" in resp.get_json()["error"]


def test_analyze_stale_data_file_400(client, monkeypatch):
    """读取时文件已消失（TOCTOU 窗口）→ 400 过期提示，非裸 500（app.py:362-366）。

    exists() 检查与 read_parquet 之间的竞态窗口无法确定性构造，
    以 read_parquet 抛 FileNotFoundError 模拟「检查后文件被清理」。"""

    def _vanished(path, *args, **kwargs):
        raise FileNotFoundError(path)

    monkeypatch.setattr(pd, "read_parquet", _vanished)
    assert _post_csv(client, "强度,温度\n45,180\n46,182\n".encode()).status_code == 200
    resp = _analyze(client)
    assert resp.status_code == 400
    assert "已过期" in resp.get_json()["error"]


def test_analyze_preprocess_error_returns_400(client):
    """特征列不存在（correlation 非 RAW_CAT 任务）→ api 层 ValidationError → 400（app.py:369-372）。"""
    assert _post_csv(client, "强度,温度\n45,180\n46,182\n".encode()).status_code == 200
    resp = _analyze(client, task="correlation", features=["不存在列"])
    assert resp.status_code == 400
    assert "数据预处理失败" in resp.get_json()["error"]


def test_analyze_unexpected_error_returns_500(client, monkeypatch):
    """非 ValidationError 的意外异常 → 500 通用中文兜底（app.py:373-375）。"""

    def _boom(*args, **kwargs):
        raise RuntimeError("模拟意外崩溃")

    monkeypatch.setattr(app_module, "run_analysis", _boom)
    assert _post_csv(client, "强度,温度\n45,180\n46,182\n".encode()).status_code == 200
    resp = _analyze(client)
    assert resp.status_code == 500
    assert "分析处理失败" in resp.get_json()["error"]


def test_analyze_power_analysis_no_data_ok(client):
    """NO_DATA 任务（power_analysis）免上传直接分析 → 200（app.py:353-356, 367-368）。"""
    resp = client.post(
        "/api/analyze",
        json={"task": "power_analysis", "targets": [], "features": [], "params": {}},
        headers={"X-CSRF-Token": _csrf(client)},
    )
    assert resp.status_code == 200, resp.get_json()
    results = resp.get_json()["results"]
    assert results and results[0]["status"] == "ok"
    assert results[0]["summary"], "power_analysis 应产出中文摘要"


# ── main() 启动入口 ──


def _silence_logging(monkeypatch):
    import smartsuite as pkg

    monkeypatch.setattr(pkg, "setup_logging", lambda: None)


def test_main_debug_forces_localhost_binding(monkeypatch, capsys):
    """debug+非本机 → 强制绑定 127.0.0.1 + 双重警告（app.py:385-395）。"""
    _silence_logging(monkeypatch)
    calls = {}
    monkeypatch.setattr(flask_app, "run", lambda **kw: calls.update(kw), raising=False)
    app_module.main(host="0.0.0.0", port=9999, debug=True)
    out = capsys.readouterr().out
    assert calls["host"] == "127.0.0.1", "debug 模式必须强制本机绑定"
    assert calls["port"] == 9999 and calls["debug"] is True
    assert "已强制绑定 127.0.0.1" in out
    assert "请勿在公网环境使用" in out


def test_main_default_runs_localhost(monkeypatch, capsys):
    """默认参数：host/port/debug 原样传递（app.py:394-398）。"""
    _silence_logging(monkeypatch)
    calls = {}
    monkeypatch.setattr(flask_app, "run", lambda **kw: calls.update(kw))
    app_module.main()
    assert calls == {"host": "127.0.0.1", "port": 5050, "debug": False}
    assert "Ctrl+C" in capsys.readouterr().out


# ── 清理链 OSError 防御分支 ──


def test_cleanup_uploads_survives_unlink_oserror(monkeypatch):
    """unlink 失败（文件被占用等）→ 静默跳过，不中断清理（app.py:56-60）。"""

    def _busy(path, *args, **kwargs):
        raise OSError("文件被占用")

    monkeypatch.setattr(os, "unlink", _busy)
    monkeypatch.setattr(os.path, "exists", lambda p: True)  # 强制进入 unlink 分支
    with app_module._upload_lock:
        app_module._UPLOAD_FILES.append("Z:/occupied/dummy.parquet")
    app_module._cleanup_uploads()  # 不应抛异常
    with app_module._upload_lock:
        app_module._UPLOAD_FILES.clear()


def test_periodic_cleanup_survives_getmtime_oserror(client, monkeypatch):
    """getmtime 失败 → 跳过该文件继续清理其余（app.py:87-100）。"""
    monkeypatch.setattr(app_module, "_CLEANUP_INTERVAL", 1)

    def _boom(path):
        if path.endswith(".parquet"):
            raise OSError("stat 失败")
        return time.time()

    real_exists = os.path.exists
    monkeypatch.setattr(os.path, "getmtime", _boom)
    stale = tempfile.NamedTemporaryFile(suffix=".parquet", delete=False)  # noqa: SIM115
    stale.close()
    with app_module._upload_lock:
        app_module._UPLOAD_FILES.append(stale.name)
    assert _post_csv(client, b"a,b\n1,2\n").status_code == 200
    # getmtime 抛 OSError → 该路径未被删除也未被移出（防御分支不中断请求）
    assert os.path.exists(stale.name)


# ── 上传解析防御分支 ──


def test_upload_csv_all_encodings_fail(client, monkeypatch):
    """全部编码均解码失败 → 「无法识别 CSV 文件编码」（app.py:249-250 防御兜底）。"""

    def _undecodable(*args, **kwargs):
        raise UnicodeDecodeError("utf-8", b"", 0, 1, "bad")

    monkeypatch.setattr(pd, "read_csv", _undecodable)
    resp = _post_csv(client, b"a,b\n1,2\n")
    assert resp.status_code == 400
    assert "无法识别 CSV 文件编码" in resp.get_json()["error"]


def test_upload_excel_zip_oversize_rejected(client):
    """zip 解压后总量 >200MB → 400（app.py:254-257 防炸弹；高压缩比数据即可触发）。"""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for i in range(110):
            zf.writestr(f"e{i}.bin", b"0" * (2 * 1024 * 1024))  # 声明大小 220MB
    resp = _post_csv(client, buf.getvalue(), filename="big.xlsx")
    assert resp.status_code == 400
    assert "解压后过大" in resp.get_json()["error"]


def test_upload_excel_over_max_rows_rejected(client):
    """xlsx 100_001 行 → 400 行数超限（app.py:275-278；CSV 路径由 test_upload_limits 覆盖）。"""
    import openpyxl

    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()
    # +2：pandas read_excel 默认首行作表头（100_002 行 → 100_001 数据行）
    for i in range(100_002):
        ws.append([float(i)])
    buf = io.BytesIO()
    wb.save(buf)
    resp = _post_csv(client, buf.getvalue(), filename="long.xlsx")
    assert resp.status_code == 400
    assert "超过限制" in resp.get_json()["error"]


def test_upload_parquet_failure_cleans_tmp_without_crash(client, monkeypatch):
    """parquet 保存失败且临时文件删除也失败 → 双重兜底后 500（app.py:295-301）。"""

    def _boom(self, *args, **kwargs):
        raise OSError("disk full")

    def _busy(path, *args, **kwargs):
        raise OSError("占用")

    monkeypatch.setattr(pd.DataFrame, "to_parquet", _boom)
    monkeypatch.setattr(os, "unlink", _busy)
    resp = _post_csv(client, b"a,b\n1,2\n")
    assert resp.status_code == 500


def test_upload_old_file_unlink_failure_keeps_new_upload(client, monkeypatch):
    """旧文件删除失败 → 跳过但新上传仍成功（app.py:306-312）。"""
    assert _post_csv(client, b"a,b\n1,2\n").status_code == 200
    with client.session_transaction() as sess:
        old_path = sess["_data_path"]

    real_unlink = os.unlink

    def _busy_first(path, *args, **kwargs):
        if path == old_path:
            raise OSError("旧文件被占用")
        return real_unlink(path)

    monkeypatch.setattr(os, "unlink", _busy_first)
    resp = _post_csv(client, b"a,b\n3,4\n")
    assert resp.status_code == 200, "旧文件删除失败不得影响新上传"
    assert os.path.exists(old_path), "被占用的旧文件保留（防御分支）"


# ── 模块入口与导入期分支（reload 舞蹈，置于文件末尾防状态扩散）──


def test_module_main_entrypoint(monkeypatch):
    """`python app.py` 入口：argparse 解析 → main() 传参（app.py:401-415）。"""
    import pathlib
    import runpy

    _silence_logging(monkeypatch)
    calls = {}
    monkeypatch.setattr("flask.Flask.run", lambda self, **kw: calls.update(kw))
    monkeypatch.setattr(sys, "argv", ["app.py", "--port", "5059"])
    runpy.run_path(str(pathlib.Path(app_module.__file__)), run_name="__main__")
    assert calls["port"] == 5059
    assert calls["host"] == "127.0.0.1"
    assert calls["debug"] is False


def test_secret_key_bootstrap_branches(monkeypatch, tmp_path, caplog):
    """SECRET_KEY 引导四分支：env 直用 / 文件读取 / 空文件重建 / 新建 + chmod 失败兜底。

    app.py:129-154 为导入期代码，通过带环境补丁的 importlib.reload 逐分支执行。
    reload 原地重执行模块对象（其他测试模块持有的引用不受影响），测试末尾
    无补丁 reload 恢复正常状态。
    """
    import importlib

    secret_file = tmp_path / ".smartsuite" / "secret_key"
    monkeypatch.setattr("pathlib.Path.home", lambda: tmp_path)

    # 分支1：env 变量直用（130-131）
    monkeypatch.setenv("SMARTSUITE_SECRET", "env-secret-abc")
    importlib.reload(app_module)
    assert app_module.app.config["SECRET_KEY"] == "env-secret-abc"
    monkeypatch.delenv("SMARTSUITE_SECRET")

    # 分支2：文件存在且有内容 → 读用（136-138, 141）
    secret_file.parent.mkdir(parents=True)
    secret_file.write_text("file-key-123", encoding="utf-8")
    importlib.reload(app_module)
    assert app_module.app.config["SECRET_KEY"] == "file-key-123"

    # 分支3：文件存在但为空 → 重建写入（138-140）
    secret_file.write_text("", encoding="utf-8")
    importlib.reload(app_module)
    assert app_module.app.config["SECRET_KEY"], "空文件应重建密钥"
    assert secret_file.read_text(encoding="utf-8") == app_module.app.config["SECRET_KEY"]

    # 分支4：文件不存在 → 新建（142-145）
    secret_file.unlink()
    importlib.reload(app_module)
    assert secret_file.exists(), "应持久化新密钥"
    assert app_module.app.config["SECRET_KEY"] == secret_file.read_text(encoding="utf-8")

    # 分支5：chmod 失败 → 忽略权限设置，密钥仍生效（147-150）
    monkeypatch.setattr(os, "chmod", lambda *a, **k: (_ for _ in ()).throw(OSError("no chown")))
    importlib.reload(app_module)
    assert app_module.app.config["SECRET_KEY"]
    monkeypatch.undo()

    # 恢复：无补丁状态重载，回读真实密钥文件
    importlib.reload(app_module)
    assert app_module.app.config["SECRET_KEY"]


def test_secret_key_home_unwritable_falls_back(monkeypatch, tmp_path, caplog):
    """home 不可写（mkdir 失败）→ 临时密钥 + 警告日志（app.py:151-154）。"""
    import importlib

    blocker = tmp_path / "not-a-dir"
    blocker.write_text("file")
    monkeypatch.setattr("pathlib.Path.home", lambda: blocker)
    with caplog.at_level(logging.WARNING, logger="smartsuite.web.app"):
        importlib.reload(app_module)
    assert app_module.app.config["SECRET_KEY"], "mkdir 失败应回退临时密钥"
    assert any("无法持久化密钥" in r.message for r in caplog.records)
    monkeypatch.undo()
    importlib.reload(app_module)  # 恢复


def test_flask_import_error_prints_guidance_and_exits(monkeypatch, capsys):
    """无 Flask 环境 → 中文安装引导 + exit(1)（app.py:19-28）。"""
    import importlib

    monkeypatch.setitem(sys.modules, "flask", None)  # import 即触发 ImportError
    with pytest.raises(SystemExit) as ei:
        importlib.reload(app_module)
    assert ei.value.code == 1
    out = capsys.readouterr().out
    assert "需要 Flask" in out and "pip install smartsuite[web]" in out
    monkeypatch.undo()
    importlib.reload(app_module)  # 恢复正常模块状态
