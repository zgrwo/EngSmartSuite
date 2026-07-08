"""过程综合审计 — 一站式工艺健康检查报告。"""
import logging
import math

import matplotlib.pyplot as plt
import numpy as np
import pandas as pd

from smartsuite.core.contracts import AnalysisRequest
from smartsuite.engine._constants import DW_SAFE_LOWER, DW_SAFE_UPPER
from smartsuite.engine._palette import PALETTE
from smartsuite.services.data_io import missing_pattern_analysis, recommend_analysis
from smartsuite.services.orchestrator import orchestrate

logger = logging.getLogger(__name__)


def _close_figures(result):
    """关闭 AnalysisResult 中的所有 matplotlib Figure，防止内存泄漏。"""
    if hasattr(result, 'figures'):
        for fig in result.figures:
            plt.close(fig)


def process_audit(
    df,
    target_col: str,
    feature_cols: list[str],
    usl: float | None = None,
    lsl: float | None = None,
    target: float | None = None,
    time_order: bool = False,
    categorical_cols: list[str] | None = None,
) -> dict:
    """过程综合审计 — 运行全套分析并生成健康评分报告。

    自动执行：数据质量 → 要因分析 → 过程能力 → 异常检测 → 趋势分析。
    返回结构化结果和建议列表。
    """
    results: dict[str, dict] = {}
    health_checks: list[dict] = []
    numeric_features = [
        c for c in feature_cols
        if pd.api.types.is_numeric_dtype(df[c])
    ]

    # ── 1. 数据质量 ──
    missing_pct = df[feature_cols + [target_col]].isna().mean().max() * 100
    if missing_pct > 5:
        health_checks.append({"检查项": "数据完整性", "状态": "⚠ 警告",
                             "详情": f"最高缺失率 {missing_pct:.1f}%"})
    else:
        health_checks.append({"检查项": "数据完整性", "状态": "✓ 正常",
                             "详情": "缺失率 < 5%"})

    # ── 2. 相关性 ──
    if len(numeric_features) >= 2:
        try:
            req = AnalysisRequest(task="correlation", data=df,
                target_col=target_col, feature_cols=numeric_features)
            r = orchestrate(req)
            _close_figures(r)
            results["correlation"] = {"status": r.status, "summary": r.summary}
            target_corr = r.metadata.get("target_correlations", {})
            corr_values = list(target_corr.values())
            # 安全提取第一个数值型相关值，非数值视为 0
            first_val = corr_values[0] if corr_values else 0
            top_r = abs(first_val) if isinstance(first_val, (int, float)) and math.isfinite(first_val) else 0
            if top_r > 0.5:
                health_checks.append({"检查项": "关键因子识别", "状态": "✓ 良好",
                                     "详情": f"存在 |r|>{top_r:.2f} 的相关因子"})
            else:
                health_checks.append({"检查项": "关键因子识别", "状态": "⚠ 注意",
                                     "详情": "无强相关因子 (|r|≤0.5)"})
        except Exception as e:
            logger.warning("关键因子识别失败", exc_info=True)
            health_checks.append({"检查项": "关键因子识别", "状态": "✗ 失败",
                                 "详情": f"计算异常 ({type(e).__name__})"})

    # ── 3. 回归/VIF ──
    if len(numeric_features) >= 3:
        try:
            req = AnalysisRequest(task="vif", data=df,
                target_col=target_col, feature_cols=numeric_features)
            r = orchestrate(req)
            _close_figures(r)
            results["vif"] = {"status": r.status, "summary": r.summary}
            high_vif = r.metadata.get("high_vif_count", 0)
            if high_vif == 0:
                health_checks.append({"检查项": "共线性诊断", "状态": "✓ 正常",
                                     "详情": "VIF 均 ≤ 5"})
            else:
                health_checks.append({"检查项": "共线性诊断", "状态": "⚠ 警告",
                                     "详情": f"{high_vif} 个因子 VIF>5"})
        except Exception as e:
            logger.warning("共线性诊断失败", exc_info=True)
            health_checks.append({"检查项": "共线性诊断", "状态": "✗ 失败",
                                 "详情": f"计算异常 ({type(e).__name__})"})

    # ── 4. 过程能力 ──
    if usl is not None and lsl is not None:
        try:
            req = AnalysisRequest(task="process_capability", data=df,
                target_col=target_col, params={"usl": usl, "lsl": lsl, "target": target})
            r = orchestrate(req)
            _close_figures(r)
            results["capability"] = {"status": r.status, "summary": r.summary}
            cpk = r.metadata.get("cpk")
            if cpk is not None and cpk >= 1.33:
                health_checks.append({"检查项": "过程能力", "状态": "✓ 合格",
                                     "详情": f"Cpk={cpk:.3f} ≥ 1.33"})
            elif cpk is not None and cpk >= 1.0:
                health_checks.append({"检查项": "过程能力", "状态": "⚠ 勉强",
                                     "详情": f"Cpk={cpk:.3f} (1.0~1.33)"})
            elif cpk is not None:
                health_checks.append({"检查项": "过程能力", "状态": "✗ 不合格",
                                     "详情": f"Cpk={cpk:.3f} < 1.0"})
            else:
                health_checks.append({"检查项": "过程能力", "状态": "—",
                                     "详情": "未计算"})
        except Exception as e:
            logger.warning("过程能力分析失败", exc_info=True)
            health_checks.append({"检查项": "过程能力", "状态": "✗ 失败",
                                 "详情": f"计算异常 ({type(e).__name__})"})

    # ── 5. 趋势 (时序数据) ──
    if time_order and len(df) >= 10:
        try:
            req = AnalysisRequest(task="trend_forecast", data=df,
                target_col=target_col, params={"forecast_steps": 5})
            r = orchestrate(req)
            _close_figures(r)
            results["trend"] = {"status": r.status, "summary": r.summary}
            dw = r.metadata.get("durbin_watson", 2)
            if DW_SAFE_LOWER <= dw <= DW_SAFE_UPPER:
                health_checks.append({"检查项": "过程稳定性", "状态": "✓ 稳定",
                                     "详情": f"DW={dw:.3f} (无自相关)"})
            else:
                health_checks.append({"检查项": "过程稳定性", "状态": "⚠ 注意",
                                     "详情": f"DW={dw:.3f} (存在自相关或趋势)"})
        except Exception as e:
            logger.warning("过程稳定性分析失败", exc_info=True)
            health_checks.append({"检查项": "过程稳定性", "状态": "✗ 失败",
                                 "详情": f"计算异常 ({type(e).__name__})"})

    # ── 6. 异常检测 ──
    try:
        req = AnalysisRequest(task="outlier_consensus", data=df,
            target_col=target_col, feature_cols=numeric_features[:3])
        r = orchestrate(req)
        _close_figures(r)
        results["outliers"] = {"status": r.status, "summary": r.summary}
        high_conf = r.metadata.get("high_confidence_count", 0)
        if high_conf == 0:
            health_checks.append({"检查项": "异常值检测", "状态": "✓ 正常",
                                 "详情": "未发现高置信异常"})
        else:
            health_checks.append({"检查项": "异常值检测", "状态": "⚠ 注意",
                                 "详情": f"{high_conf} 个高置信异常点"})
    except Exception as e:
        logger.warning("异常值检测失败", exc_info=True)
        health_checks.append({"检查项": "异常值检测", "状态": "✗ 失败",
                             "详情": f"计算异常 ({type(e).__name__})"})

    # ── 汇总评分 ──
    ok_count = sum(1 for h in health_checks if "✓" in str(h["状态"]))
    warn_count = sum(1 for h in health_checks if "⚠" in str(h["状态"]))
    fail_count = sum(1 for h in health_checks if "✗" in str(h["状态"]))
    total = len(health_checks)

    if fail_count > 0:
        overall = "需整改 (存在不合格项)"
    elif warn_count >= 2:
        overall = "需关注 (多项警告)"
    elif warn_count == 1:
        overall = "良好 (1 项需关注)"
    else:
        overall = "优秀 (全部正常)"

    return {
        "health_checks": pd.DataFrame(health_checks),
        "results_summary": {k: v.get("summary", "") for k, v in results.items()},
        "overall_rating": overall,
        "score_detail": f"✓ {ok_count} / ⚠ {warn_count} / ✗ {fail_count} (共{total}项)",
        "summary": f"过程审计: {overall} ({ok_count}/{total} 项正常)",
    }


def batch_analyze(df, target_col, feature_cols, tasks=None, **kwargs):
    """批量运行多个分析任务，返回结果摘要字典。"""
    if tasks is None:
        tasks = ["correlation", "regression", "vif", "anova",
                "normality_check", "distribution_summary"]

    results = {}
    for task in tasks:
        try:
            req = AnalysisRequest(task=task, data=df, target_col=target_col,
                                  feature_cols=feature_cols,
                                  params=kwargs.get(task, {}))
            r = orchestrate(req)
            _close_figures(r)
            results[task] = {"status": r.status, "summary": r.summary}
        except Exception as e:
            logger.warning("批量分析任务「%s」执行失败: %s", task, e)
            results[task] = {"status": "error",
                             "summary": f"分析任务「{task}」执行失败，请检查数据格式和参数配置。"}

    ok = sum(1 for v in results.values() if v["status"] == "ok")
    return {"results": results, "summary": f"Batch: {ok}/{len(tasks)} tasks OK"}


def auto_report(df, target_col, feature_cols=None, output_path=None,
                usl=None, lsl=None, title="SmartSuite 自动分析报告"):
    """一键自动报告 — 数据诊断 → 智能推荐 → 批量分析 → HTML 输出。

    Args:
        df: DataFrame
        target_col: 目标列名
        feature_cols: 因子列 (None=自动选择数值列)
        output_path: HTML 输出路径 (None=自动生成)
        usl/lsl: 规格限 (可选)
        title: 报告标题
    """
    import os

    from smartsuite.services.reporter import to_html

    if feature_cols is None:
        feature_cols = [c for c in df.columns
                       if pd.api.types.is_numeric_dtype(df[c]) and c != target_col]

    # 数据质量
    quality = missing_pattern_analysis(df)
    rec = recommend_analysis(df, target_col=target_col)

    # 批量分析 (不含 regression，因其将在后续 HTML 生成时单独运行)
    tasks_to_run = ["correlation", "anova", "normality_check",
                   "distribution_summary"]
    if usl is not None and lsl is not None:
        tasks_to_run.append("process_capability")
    batch = batch_analyze(df, target_col, feature_cols, tasks=tasks_to_run,
                          **{"process_capability": {"usl": usl, "lsl": lsl}})

    # 综合审计
    audit = process_audit(df, target_col=target_col, feature_cols=feature_cols,
                          usl=usl, lsl=lsl)

    # 生成 HTML
    if output_path is None:
        output_path = os.path.join(os.getcwd(), "smartsuite_report.html")

    # 使用回归分析作为代表性结果生成 HTML
    feat_for_html = feature_cols[:min(len(feature_cols), 10)]
    req = AnalysisRequest(task="regression", data=df, target_col=target_col,
                          feature_cols=feat_for_html)
    r = orchestrate(req)
    if r.status != "ok":
        r.summary = (
            f"自动分析报告: {title}\n"
            f"数据质量: {quality['summary']}\n"
            f"分析推荐: {rec['summary']}\n"
            f"批量分析: {batch['summary']}\n"
            f"综合审计: {audit['summary']}\n"
            f"---\n回归建模失败，无法生成详细报告"
        )
    else:
        r.summary = (
        f"自动分析报告: {title}\n"
        f"数据质量: {quality['summary']}\n"
        f"分析推荐: {rec['summary']}\n"
        f"批量分析: {batch['summary']}\n"
        f"综合审计: {audit['summary']}\n"
        f"---\n{r.summary}"
    )
    to_html(r, output_path)
    _close_figures(r)

    return {
        "output_path": output_path,
        "data_quality": quality,
        "recommendations": rec,
        "batch_results": batch,
        "audit": audit,
        "summary": f"报告已生成: {output_path}\n{audit['summary']}",
    }


def export_workbook(df, target_col, feature_cols, output_path, tasks=None):
    """批量分析导出到多 Sheet Excel 工作簿。

    每个分析任务的结果 (表格+图表) 写入独立 Sheet。
    """
    import os

    import openpyxl

    # 确保输出目录存在 (与 reporter.to_pdf/to_ppt/to_html 保持一致)
    out_dir = os.path.dirname(os.path.abspath(output_path))
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    if tasks is None:
        tasks = ["correlation", "regression", "anova", "distribution_summary",
                "normality_check", "vif"]

    for task in tasks:
        try:
            req = AnalysisRequest(task=task, data=df, target_col=target_col,
                                  feature_cols=feature_cols)
            r = orchestrate(req)
            _close_figures(r)
            if r.status != "ok":
                continue

            # Summary sheet
            ws = wb.create_sheet(title=f"{task}_summary"[:31])
            ws["A1"] = f"分析: {task}"
            ws["A2"] = f"结论: {r.summary}"
            ws["A1"].font = openpyxl.styles.Font(bold=True, size=12)
            row = 4

            # Tables
            for name, table in r.tables.items():
                ws.cell(row=row, column=1, value=name).font = openpyxl.styles.Font(bold=True)
                row += 1
                # Headers
                for ci, col in enumerate(table.columns):
                    ws.cell(row=row, column=ci+1, value=str(col)).font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                    ws.cell(row=row, column=ci+1).fill = openpyxl.styles.PatternFill("solid", fgColor=PALETTE["data"]["primary"])
                row += 1
                # Data
                for _, data_row in table.head(100).iterrows():
                    for ci, val in enumerate(data_row):
                        if val is None:
                            ws.cell(row=row, column=ci+1, value="")
                        elif isinstance(val, (int, float)):
                            # np.nan is a float — must guard before writing to openpyxl
                            if isinstance(val, float) and np.isnan(val):
                                ws.cell(row=row, column=ci+1, value="NaN")
                            else:
                                ws.cell(row=row, column=ci+1, value=val)
                        else:
                            ws.cell(row=row, column=ci+1, value=str(val)[:100])
                    row += 1
                row += 2

        except Exception:
            logger.warning("导出工作表失败: %s", task, exc_info=True)
            continue

    if len(wb.sheetnames) == 0:
        # 所有任务均失败，创建一个错误信息 Sheet
        ws = wb.create_sheet(title="导出状态")
        ws["A1"] = "所有分析任务均失败，无法生成报告"
        ws["A2"] = f"已尝试任务: {', '.join(tasks)}"
    wb.save(output_path)
    return output_path
